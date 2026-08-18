import logging
from io import BytesIO
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)


class ExcelImportService:

    # 常见中文优先级到内部缺陷级别的映射。
    SEVERITY_MAP = {
        "致命": "critical", "严重": "major", "高": "major",
        "一般": "minor", "中": "minor",
        "轻微": "trivial", "低": "trivial",
        "critical": "critical", "major": "major", "minor": "minor", "trivial": "trivial",
        "紧急": "critical", "高优先级": "major",
    }

    @classmethod
    def _map_severity(cls, value: str) -> str:
        v = str(value).strip()
        return cls.SEVERITY_MAP.get(v, "minor")

    @classmethod
    def parse_defect_records(cls, file_content: bytes) -> list[dict[str, Any]]:
        """从缺陷 Excel 导入记录，识别标题、描述、优先级列并忽略多余列。"""
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # 建立表头到列下标的映射。
        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = {h: i for i, h in enumerate(headers)}

        # 必填列。
        title_col = col_map.get("标题")
        desc_col = col_map.get("描述")
        severity_col = col_map.get("优先级")

        if title_col is None and desc_col is None:
            raise ValueError("Excel 缺少必要列：标题 和 描述")

        result = []
        for row in rows[1:]:
            title = str(row[title_col]).strip() if title_col is not None and row[title_col] else ""
            desc = str(row[desc_col]).strip() if desc_col is not None and row[desc_col] else ""
            if not title and not desc:
                continue

            severity = "minor"
            if severity_col is not None and row[severity_col]:
                severity = cls._map_severity(str(row[severity_col]))

            result.append({"title": title, "description": desc, "severity": severity})

        wb.close()
        return result


class ExcelExportService:
    """导出测试用例到 Excel。"""

    @classmethod
    def export_test_cases(cls, cases: list[dict[str, Any]]) -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试用例"
        ws.append(["用例标题", "等级", "前置条件", "步骤描述", "预期结果"])
        for c in cases:
            steps_str = c.get("steps", "")
            if isinstance(steps_str, list):
                steps_str = "\n".join(
                    f"{s.get('step_no', i+1)}. {s.get('action', str(s))}" if isinstance(s, dict) else str(s)
                    for i, s in enumerate(steps_str)
                )
            elif isinstance(steps_str, str):
                try:
                    import json
                    parsed = json.loads(steps_str)
                    if isinstance(parsed, list):
                        steps_str = "\n".join(f"{s.get('step_no', i+1)}. {s.get('action', str(s))}" if isinstance(s, dict) else str(s) for i, s in enumerate(parsed))
                except Exception:
                    logger.debug("步骤字段不是 JSON 数组，按原始字符串导出")

            ws.append([
                c.get("title", ""),
                c.get("priority", "P1"),
                c.get("precondition", ""),
                steps_str,
                c.get("expected_result", ""),
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        return output
